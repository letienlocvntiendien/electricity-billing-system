#!/usr/bin/env python3
"""
generate_historical_migration.py

Reads docs/data/clean_data_migration_workbook2020_2026.xlsx and generates
src/main/resources/db/migration/V10__seed_historical_data.sql

Usage:
    python3 scripts/generate_historical_migration.py

Requirements:
    pip install openpyxl
"""

import os
import calendar
from datetime import date, datetime
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: python3 -m pip install openpyxl")
    raise

# ─────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

EXCEL_PATH = os.path.join(
    BASE_DIR, "docs", "data", "clean_data_migration_workbook2020_2026.xlsx"
)

OUTPUT_SQL_PATH = os.path.join(
    BASE_DIR, "src", "main", "resources", "db", "migration",
    "V10__seed_historical_data.sql"
)

# KH093 "HĐ" is not a customer (government entity), exclude entirely
EXCLUDED_CUSTOMERS = {"KH093"}

# Flags in need_review that make a row un-importable
SKIP_FLAGS = {
    "MULTI_MONTH_ROW",    # belongs to multi-month sheet, period ambiguous
    "MISSING_INDEX",      # currentIndex is NULL → violates NOT NULL
    "INDEX_REVERSED",     # currentIndex < previousIndex → violates CHECK constraint
    "NEGATIVE_CONSUMPTION",  # same issue as INDEX_REVERSED
    "NEGATIVE_AMOUNT",    # do not auto-import negative billing adjustments
}


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def sql_str(value):
    """Escape a string for SQL single-quoted literal."""
    if value is None:
        return "NULL"
    return "'" + str(value).replace("\\", "\\\\").replace("'", "\\'") + "'"


def period_start_end(period_code: str):
    """'2021-01' → (date(2021,1,1), date(2021,1,31))"""
    year = int(period_code[:4])
    month = int(period_code[5:7])
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def is_valid_period_code(code) -> bool:
    if not code or "_to_" in str(code):
        return False
    s = str(code)
    if len(s) != 7 or s[4] != "-":
        return False
    try:
        year = int(s[:4])
        month = int(s[5:7])
        return 2000 <= year <= 2030 and 1 <= month <= 12
    except ValueError:
        return False


def mode_of(values):
    """Return most common non-None value, or None if list is empty."""
    filtered = [v for v in values if v is not None]
    if not filtered:
        return None
    return Counter(filtered).most_common(1)[0][0]


def to_int(value, default=0) -> int:
    if value is None:
        return default
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return default


def to_float(value, default=0.0) -> float:
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def round0(value) -> int:
    """Round to nearest integer (0 decimal places) — for money amounts."""
    return round(to_float(value))


# ─────────────────────────────────────────────────────────────
# Step 1: Read Excel sheets
# ─────────────────────────────────────────────────────────────

print("Reading Excel file …")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# ── customers sheet ──
# cols: provisional_code(0), name_normalized(1), ...
ws_customers = wb["customers"]
customer_rows = list(ws_customers.iter_rows(values_only=True))[1:]  # skip header
customer_names: dict[str, str] = {}
for r in customer_rows:
    code = r[0]
    name = r[1]
    if code and code not in EXCLUDED_CUSTOMERS:
        customer_names[code] = name or code

print(f"  customers sheet: {len(customer_names)} entries (excluded: {EXCLUDED_CUSTOMERS})")

# ── billing_periods sheet ──
# cols: period_code(0), sheet_name(1), period_start(2), period_end(3),
#       unit_price(4), service_fee(5), num_customers(6), is_multi_month(7), flags(8)
ws_bp = wb["billing_periods"]
bp_rows = list(ws_bp.iter_rows(values_only=True))[1:]
billing_periods_meta: dict[str, dict] = {}
for r in bp_rows:
    code = r[0]
    is_multi = r[7]
    if not is_valid_period_code(code) or is_multi:
        continue
    billing_periods_meta[code] = {
        "unit_price": r[4],
        "service_fee": r[5],
    }

print(f"  billing_periods sheet: {len(billing_periods_meta)} single-month periods")

# ── import_candidates sheet ──
# cols: period_code(0), customer_provisional_code(1), customer_raw_name(2),
#       source_sheet(3), excel_row(4), previous_index(5), current_index(6),
#       consumption_excel(7), consumption_calculated(8), unit_price(9),
#       electricity_amount(10), service_fee(11), total_amount(12),
#       is_multi_month(13), paid_marker(14), flags(15), VERIFY(16)
ws_ic = wb["import_candidates"]
ic_rows = list(ws_ic.iter_rows(values_only=True))[1:]
import_candidates = [r for r in ic_rows if r[0] and r[1]]
print(f"  import_candidates sheet: {len(import_candidates)} rows")

# ── need_review sheet ──
# same columns as import_candidates (0-15), no VERIFY column
ws_nr = wb["need_review"]
nr_rows = list(ws_nr.iter_rows(values_only=True))[1:]

need_review_valid = []
nr_skipped_by_flag: dict[str, int] = Counter()
for r in nr_rows:
    if not r[0] or not r[1]:
        continue
    flags_str = str(r[15]) if r[15] else ""
    row_flags = {f.strip() for f in flags_str.split(";")} if flags_str else set()
    skipped_flags = row_flags & SKIP_FLAGS
    if skipped_flags:
        for f in skipped_flags:
            nr_skipped_by_flag[f] += 1
        continue
    need_review_valid.append(r)

print(f"  need_review sheet: {len(need_review_valid)} importable rows "
      f"(skipped: {dict(nr_skipped_by_flag)})")

wb.close()

# ─────────────────────────────────────────────────────────────
# Step 2: Merge, exclude KH093, dedup
# ─────────────────────────────────────────────────────────────

all_rows = import_candidates + need_review_valid

# Exclude KH093
before_excl = len(all_rows)
all_rows = [r for r in all_rows if r[1] not in EXCLUDED_CUSTOMERS]
kh093_excluded = before_excl - len(all_rows)

# Also skip rows with invalid period code
all_rows = [r for r in all_rows if is_valid_period_code(r[0])]

# Dedup by (period_code, customer_code)
# Keep row with MAX(consumption_excel); on tie keep MIN(excel_row)
grouped: dict[tuple, list] = defaultdict(list)
for r in all_rows:
    grouped[(r[0], r[1])].append(r)

deduplicated = []
phantom_removed = 0
for (period_code, customer_code), group in grouped.items():
    if len(group) == 1:
        deduplicated.append(group[0])
    else:
        def sort_key(r):
            cons = to_float(r[7], -1)
            row_num = to_int(r[4], 9999)
            return (-cons, row_num)   # descending consumption, ascending row number
        group.sort(key=sort_key)
        deduplicated.append(group[0])
        phantom_removed += len(group) - 1

all_rows = deduplicated
print(f"\n  KH093 excluded:       {kh093_excluded}")
print(f"  Phantom rows removed: {phantom_removed}")
print(f"  Rows after dedup:     {len(all_rows)}")

# ─────────────────────────────────────────────────────────────
# Step 3: Fill missing fields, calculate totals
# ─────────────────────────────────────────────────────────────

# Collect per-period mode values from all rows with data
period_unit_prices: dict[str, list] = defaultdict(list)
period_service_fees: dict[str, list] = defaultdict(list)
for r in all_rows:
    if r[9] is not None:
        period_unit_prices[r[0]].append(to_float(r[9]))
    if r[11] is not None:
        period_service_fees[r[0]].append(to_float(r[11]))

period_mode_up = {p: mode_of(ups) for p, ups in period_unit_prices.items()}
period_mode_sf = {p: mode_of(sfs) for p, sfs in period_service_fees.items()}

null_total_filled = 0
null_up_filled = 0
processed_rows = []

for r in all_rows:
    r = list(r)   # make mutable
    period_code: str = r[0]
    customer_code: str = r[1]

    # Fill unit_price
    unit_price = to_float(r[9]) if r[9] is not None else None
    if unit_price is None:
        unit_price = period_mode_up.get(period_code)
        if unit_price is not None:
            null_up_filled += 1
    unit_price = unit_price or 0.0
    r[9] = unit_price

    # Fill service_fee (default 20000 if completely unknown)
    service_fee = to_float(r[11]) if r[11] is not None else None
    if service_fee is None:
        service_fee = period_mode_sf.get(period_code, 20000.0)
    r[11] = service_fee

    # Consumption
    consumption = to_int(r[7]) if r[7] is not None else 0
    r[7] = consumption

    # electricity_amount
    elec_amount = round0(r[10]) if r[10] is not None else round(consumption * unit_price)
    r[10] = elec_amount

    # total_amount
    total_amount = r[12]
    if total_amount is None:
        total_amount = elec_amount + round(service_fee)
        null_total_filled += 1
    r[12] = round0(total_amount)

    # previous_index / current_index
    prev_idx = to_int(r[5], 0)
    curr_idx = to_int(r[6], prev_idx)   # if None, same as prev (0 consumption)
    if curr_idx < prev_idx:
        # Safety net — should already be excluded but skip if not
        print(f"  SKIP INDEX_REVERSED: {period_code} {customer_code} "
              f"prev={prev_idx} curr={curr_idx}")
        continue
    r[5] = prev_idx
    r[6] = curr_idx

    processed_rows.append(r)

print(f"  NULL unit_price filled:   {null_up_filled}")
print(f"  NULL total_amount filled: {null_total_filled}")

# ─────────────────────────────────────────────────────────────
# Step 4: Aggregate period-level fields
# ─────────────────────────────────────────────────────────────

period_agg: dict[str, dict] = defaultdict(lambda: {
    "unit_prices": [], "service_fees": [], "total_kwh": 0
})
for r in processed_rows:
    p = period_agg[r[0]]
    p["unit_prices"].append(r[9])
    p["service_fees"].append(r[11])
    if r[7] and r[7] > 0:
        p["total_kwh"] += r[7]

periods_final: dict[str, dict] = {}
for code, agg in period_agg.items():
    if not is_valid_period_code(code):
        continue
    unit_price = mode_of(agg["unit_prices"]) or 0.0
    service_fee = mode_of(agg["service_fees"]) or 20000.0
    total_kwh = agg["total_kwh"]
    total_amount = round(unit_price * total_kwh, 2)
    start_dt, end_dt = period_start_end(code)
    year = int(code[:4])
    month = int(code[5:7])
    periods_final[code] = {
        "code": code,
        "name": f"Kỳ tháng {month:02d}/{year}",
        "start_date": start_dt,
        "end_date": end_dt,
        "unit_price": unit_price,
        "service_fee": service_fee,
        "evn_total_kwh": total_kwh,
        "evn_total_amount": total_amount,
    }

# Customers that actually appear in data rows (only import these)
customers_in_data = set(r[1] for r in processed_rows)
customers_to_import = {
    code: customer_names.get(code, code)
    for code in customers_in_data
}

print(f"  Periods to insert:   {len(periods_final)}")
print(f"  Customers to insert: {len(customers_to_import)}")

# ─────────────────────────────────────────────────────────────
# Step 5: Generate SQL
# ─────────────────────────────────────────────────────────────

lines = []

def section(title, count=""):
    lines.append("")
    lines.append("-- " + "─" * 60)
    lines.append(f"-- {title}" + (f"  ({count} rows)" if count else ""))
    lines.append("-- " + "─" * 60)


lines += [
    "-- " + "═" * 60,
    "-- V10__seed_historical_data.sql",
    "-- Generated by scripts/generate_historical_migration.py",
    f"-- Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    f"-- Customers:      {len(customers_to_import)}",
    f"-- Billing periods:{len(periods_final)}",
    f"-- Meter readings: {len(processed_rows)}",
    f"-- Bills:          {len(processed_rows)}",
    "-- DO NOT EDIT MANUALLY",
    "-- " + "═" * 60,
    "",
    "SET NAMES utf8mb4;",
    "SET FOREIGN_KEY_CHECKS = 0;",
]

# ── CUSTOMERS ──────────────────────────────────────────────
section("CUSTOMERS", len(customers_to_import))
for code in sorted(customers_to_import.keys()):
    name = customers_to_import[code]
    lines.append(
        f"INSERT IGNORE INTO customer (code, full_name, active, created_at, updated_at) "
        f"VALUES ({sql_str(code)}, {sql_str(name)}, TRUE, NOW(), NOW());"
    )

# ── BILLING PERIODS ────────────────────────────────────────
section("BILLING PERIODS", len(periods_final))
for code in sorted(periods_final.keys()):
    p = periods_final[code]
    sd = p["start_date"].strftime("%Y-%m-%d")
    ed = p["end_date"].strftime("%Y-%m-%d")
    closed_at = f"{ed} 23:59:59"
    lines.append(
        f"INSERT IGNORE INTO billing_period "
        f"(code, name, start_date, end_date, unit_price, service_fee, extra_fee, "
        f"evn_total_kwh, evn_total_amount, status, closed_at, created_at, updated_at) VALUES ("
        f"{sql_str(code)}, {sql_str(p['name'])}, '{sd}', '{ed}', "
        f"{p['unit_price']}, {p['service_fee']}, 0, "
        f"{p['evn_total_kwh']}, {p['evn_total_amount']}, "
        f"'CLOSED', '{closed_at}', NOW(), NOW());"
    )

# ── EVN INVOICES ───────────────────────────────────────────
section("EVN INVOICES (1 synthetic per period)", len(periods_final))
lines.append(
    "-- Each period gets one synthetic invoice derived from total consumption."
)
lines.append(
    "-- Uses INSERT...SELECT so period_id is resolved correctly even if id is not known."
)
for code in sorted(periods_final.keys()):
    p = periods_final[code]
    ed = p["end_date"].strftime("%Y-%m-%d")
    inv_num = f"HIST-{code}"
    kwh = p["evn_total_kwh"]
    amount = p["evn_total_amount"]
    lines.append(
        f"INSERT INTO evn_invoice (period_id, invoice_date, invoice_number, kwh, amount, created_at) "
        f"SELECT bp.id, '{ed}', {sql_str(inv_num)}, {kwh}, {amount}, NOW() "
        f"FROM billing_period bp "
        f"WHERE bp.code = {sql_str(code)} "
        f"AND NOT EXISTS (SELECT 1 FROM evn_invoice ei WHERE ei.period_id = bp.id);"
    )

# ── METER READINGS ─────────────────────────────────────────
section("METER READINGS", len(processed_rows))
lines.append(
    "-- consumption column is STORED GENERATED (currentIndex - previousIndex) — NOT inserted."
)
for r in processed_rows:
    period_code = r[0]
    customer_code = r[1]
    prev_idx = r[5]
    curr_idx = r[6]
    p = periods_final.get(period_code)
    if p is None:
        continue
    read_at = p["end_date"].strftime("%Y-%m-%d") + " 12:00:00"
    lines.append(
        f"INSERT IGNORE INTO meter_reading "
        f"(period_id, customer_id, previous_index, current_index, read_at, created_at, updated_at) "
        f"SELECT bp.id, c.id, {prev_idx}, {curr_idx}, "
        f"'{read_at}', NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(period_code)} AND c.code = {sql_str(customer_code)};"
    )

# ── BILLS ──────────────────────────────────────────────────
section("BILLS", len(processed_rows))
lines.append(
    "-- consumption = consumption_excel (what was actually billed, matches Excel)."
)
lines.append(
    "-- status = PAID, paid_amount = total_amount (all historical bills treated as paid)."
)
for r in processed_rows:
    period_code = r[0]
    customer_code = r[1]
    consumption = r[7]
    unit_price = r[9]
    service_fee = r[11]
    elec_amount = r[10]
    service_amount = round(service_fee)
    total_amount = r[12]
    paid_amount = total_amount
    payment_code = f"TIENDIEN {period_code} {customer_code}"

    p = periods_final.get(period_code)
    if p is None:
        continue

    lines.append(
        f"INSERT IGNORE INTO bill "
        f"(period_id, customer_id, consumption, unit_price, service_fee, "
        f"electricity_amount, service_amount, total_amount, paid_amount, "
        f"status, payment_code, sent_via_zalo, created_at, updated_at) "
        f"SELECT bp.id, c.id, {consumption}, {unit_price}, {service_fee}, "
        f"{elec_amount}, {service_amount}, {total_amount}, {paid_amount}, "
        f"'PAID', {sql_str(payment_code)}, FALSE, NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(period_code)} AND c.code = {sql_str(customer_code)};"
    )

lines += ["", "SET FOREIGN_KEY_CHECKS = 1;", ""]

# ─────────────────────────────────────────────────────────────
# Step 6: Write output
# ─────────────────────────────────────────────────────────────

sql_content = "\n".join(lines)
os.makedirs(os.path.dirname(OUTPUT_SQL_PATH), exist_ok=True)
with open(OUTPUT_SQL_PATH, "w", encoding="utf-8") as f:
    f.write(sql_content)

# ─────────────────────────────────────────────────────────────
# Final summary
# ─────────────────────────────────────────────────────────────

print()
print("=" * 55)
print("  Migration Summary")
print("=" * 55)
print(f"  Customers:        {len(customers_to_import)}")
period_list = sorted(periods_final.keys())
print(f"  Billing periods:  {len(periods_final)}"
      f"  ({period_list[0]} to {period_list[-1]})")
print(f"  EVN invoices:     {len(periods_final)}")
print(f"  Meter readings:   {len(processed_rows)}")
print(f"  Bills:            {len(processed_rows)}")
print()
print("  Exclusions / Fixes")
print(f"  KH093 excluded:       {kh093_excluded}")
print(f"  Phantom rows removed: {phantom_removed}")
for flag, count in sorted(nr_skipped_by_flag.items(), key=lambda x: -x[1]):
    print(f"  {flag:28s}: {count}")
print()
print("  Data fixes")
print(f"  NULL unit_price filled:    {null_up_filled}")
print(f"  NULL total_amount filled:  {null_total_filled}")
sf40 = sum(1 for r in processed_rows if r[11] == 40000)
print(f"  service_fee=40000 rows:    {sf40}  (2022-12 — intentional)")
print()
print(f"  Output: {OUTPUT_SQL_PATH}")
print(f"  File size: {len(sql_content.encode('utf-8')):,} bytes")
print("=" * 55)
