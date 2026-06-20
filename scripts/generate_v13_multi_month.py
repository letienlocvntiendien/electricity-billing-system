#!/usr/bin/env python3
"""
generate_v13_multi_month.py

Generates V13__import_multi_month_periods.sql

Imports all 316 MULTI_MONTH_ROW rows from need_review, treating each
multi-month sheet as a single billing period (gộp nhiều tháng thành 1 kỳ).

5 multi-month periods present in need_review:
  - 2019-06_to_2019-07 (59 customers)
  - 2021-10_to_2021-11 (62 customers)
  - 2022-05_to_2022-06 (62 customers, 1 meter reset)
  - 2023-06_to_2023-07 (64 customers)
  - 2026-01_to_2026-02 (69 customers)

Run AFTER V12 has been applied.

Usage:
    python3 scripts/generate_v13_multi_month.py
"""

import os, calendar
from datetime import date, datetime
from collections import defaultdict, Counter

try:
    import openpyxl
except ImportError:
    raise SystemExit("Run: python3 -m pip install openpyxl")

BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXCEL_PATH = os.path.join(BASE_DIR, "docs", "data",
             "clean_data_migration_workbook2020_2026.xlsx")
OUTPUT_PATH = os.path.join(BASE_DIR, "src", "main", "resources", "db", "migration",
              "V13__import_multi_month_periods.sql")

EXCLUDED_CUSTOMERS = {"KH093"}


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def sql_str(v):
    if v is None: return "NULL"
    return "'" + str(v).replace("\\", "\\\\").replace("'", "\\'") + "'"

def to_int(v, default=0):
    if v is None: return default
    try: return int(float(v))
    except: return default

def to_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def round0(v): return round(to_float(v))

def mode_of(values):
    filtered = [v for v in values if v is not None]
    return Counter(filtered).most_common(1)[0][0] if filtered else None

def parse_multi_period(code: str):
    """
    '2019-06_to_2019-07' → (start_date, end_date, period_name)
    """
    parts = code.split("_to_")
    start_str = parts[0]   # 'YYYY-MM'
    end_str   = parts[1]   # 'YYYY-MM'
    sy, sm = int(start_str[:4]), int(start_str[5:7])
    ey, em = int(end_str[:4]),   int(end_str[5:7])
    start_date = date(sy, sm, 1)
    last_day   = calendar.monthrange(ey, em)[1]
    end_date   = date(ey, em, last_day)
    if sy == ey:
        name = f"Kỳ tháng {sm:02d}-{em:02d}/{sy}"
    else:
        name = f"Kỳ tháng {sm:02d}/{sy} - {em:02d}/{ey}"
    return start_date, end_date, name

def is_multi_month_code(code) -> bool:
    return code and "_to_" in str(code)


# ─────────────────────────────────────────────────────────────
# Read Excel
# ─────────────────────────────────────────────────────────────

print("Reading Excel …")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

ws_nr  = wb["need_review"]
nr_data = list(ws_nr.iter_rows(values_only=True))[1:]

# Collect existing unit_prices per single-month period for fallback
ws_ic  = wb["import_candidates"]
ic_rows = list(ws_ic.iter_rows(values_only=True))[1:]
fallback_sf  = defaultdict(list)
for r in ic_rows:
    if r[11] is not None: fallback_sf[r[0]].append(to_float(r[11]))
period_sf_mode = {p: mode_of(vals) for p, vals in fallback_sf.items()}

wb.close()

multi_rows = [r for r in nr_data if r[15] and "MULTI_MONTH_ROW" in str(r[15])]
print(f"  MULTI_MONTH_ROW rows: {len(multi_rows)}")

# Group by period_code
by_period: dict[str, list] = defaultdict(list)
for r in multi_rows:
    if r[1] not in EXCLUDED_CUSTOMERS:
        by_period[r[0]].append(r)

print(f"  Unique multi-month periods: {len(by_period)}")


# ─────────────────────────────────────────────────────────────
# Process each period
# ─────────────────────────────────────────────────────────────

period_stats = []   # for summary
lines = []
total_mr = 0
total_bill = 0
skipped_rows = 0

def mr_sql(period_code, customer_code, prev_idx, curr_idx, read_at):
    return (
        f"INSERT IGNORE INTO meter_reading "
        f"(period_id, customer_id, previous_index, current_index, read_at, created_at, updated_at) "
        f"SELECT bp.id, c.id, {prev_idx}, {curr_idx}, "
        f"'{read_at}', NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(period_code)} AND c.code = {sql_str(customer_code)};"
    )

def bill_sql(period_code, customer_code, cons, up, sf, elec, total):
    svc = round(sf)
    payment_code = f"TIENDIEN {period_code} {customer_code}"
    return (
        f"INSERT IGNORE INTO bill "
        f"(period_id, customer_id, consumption, unit_price, service_fee, "
        f"electricity_amount, service_amount, total_amount, paid_amount, "
        f"status, payment_code, sent_via_zalo, created_at, updated_at) "
        f"SELECT bp.id, c.id, {cons}, {up}, {sf}, "
        f"{elec}, {svc}, {total}, {total}, "
        f"'PAID', {sql_str(payment_code)}, FALSE, NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(period_code)} AND c.code = {sql_str(customer_code)};"
    )


for period_code in sorted(by_period.keys()):
    rows = by_period[period_code]

    # Parse dates and name
    try:
        start_date, end_date, period_name = parse_multi_period(period_code)
    except Exception as e:
        print(f"  SKIP {period_code}: cannot parse dates ({e})")
        skipped_rows += len(rows)
        continue

    sd  = start_date.strftime("%Y-%m-%d")
    ed  = end_date.strftime("%Y-%m-%d")
    read_at = f"{ed} 12:00:00"

    # Period-level aggregates
    unit_prices  = [to_float(r[9]) for r in rows if r[9] is not None]
    service_fees = [to_float(r[11]) for r in rows if r[11] is not None]
    unit_price   = mode_of(unit_prices) or 0.0
    service_fee  = mode_of(service_fees) or 0.0   # 0 if no data (e.g. 2019-06_to_2019-07)
    total_kwh    = sum(to_int(r[7]) for r in rows if r[7] is not None and to_float(r[7]) > 0)
    evn_amount   = round(unit_price * total_kwh, 2)
    inv_num      = f"HIST-{period_code}"

    lines += [
        "",
        "-- " + "─" * 62,
        f"-- Period: {period_code}  ({period_name})",
        f"-- Date range: {sd} → {ed}",
        f"-- Customers: {len(rows)}  |  unit_price: {unit_price}  |  service_fee: {service_fee}",
        f"-- total_kwh: {total_kwh}  |  evn_amount: {evn_amount}",
        "-- " + "─" * 62,
    ]

    # billing_period
    lines += [
        f"INSERT IGNORE INTO billing_period "
        f"(code, name, start_date, end_date, unit_price, service_fee, extra_fee, "
        f"evn_total_kwh, evn_total_amount, status, closed_at, created_at, updated_at) VALUES ("
        f"{sql_str(period_code)}, {sql_str(period_name)}, '{sd}', '{ed}', "
        f"{unit_price}, {service_fee}, 0, "
        f"{total_kwh}, {evn_amount}, "
        f"'CLOSED', '{ed} 23:59:59', NOW(), NOW());",
    ]

    # evn_invoice
    lines += [
        f"INSERT INTO evn_invoice (period_id, invoice_date, invoice_number, kwh, amount, created_at) "
        f"SELECT bp.id, '{ed}', {sql_str(inv_num)}, {total_kwh}, {evn_amount}, NOW() "
        f"FROM billing_period bp "
        f"WHERE bp.code = {sql_str(period_code)} "
        f"AND NOT EXISTS (SELECT 1 FROM evn_invoice ei WHERE ei.period_id = bp.id);",
    ]

    # meter_readings + bills
    lines.append("-- Meter readings")
    period_mr = 0
    for r in rows:
        customer_code = r[1]
        flags = {f.strip() for f in str(r[15]).split(";")} if r[15] else set()
        cons_excel = to_int(r[7])
        up_row     = to_float(r[9]) or unit_price
        sf_row     = to_float(r[11]) if r[11] is not None else service_fee

        if "INDEX_REVERSED" in flags:
            # Meter replacement: prev=0, curr=cons_excel
            if cons_excel <= 0:
                skipped_rows += 1
                continue
            prev_idx = 0
            curr_idx = cons_excel
        elif "MISSING_INDEX" in flags:
            prev_idx = to_int(r[5], 0)
            curr_idx = to_int(r[6], prev_idx)
            if curr_idx < prev_idx:
                skipped_rows += 1
                continue
        else:
            prev_idx = to_int(r[5], 0)
            curr_idx = to_int(r[6], prev_idx)
            if curr_idx < prev_idx:
                skipped_rows += 1
                continue

        lines.append(mr_sql(period_code, customer_code, prev_idx, curr_idx, read_at))
        period_mr += 1

    lines.append("-- Bills")
    period_bill = 0
    for r in rows:
        customer_code = r[1]
        flags = {f.strip() for f in str(r[15]).split(";")} if r[15] else set()
        cons_excel = to_int(r[7])
        up_row     = to_float(r[9]) or unit_price
        sf_row     = to_float(r[11]) if r[11] is not None else service_fee
        elec       = round0(r[10]) if r[10] is not None else round(cons_excel * up_row)
        total      = round0(r[12]) if r[12] is not None else (elec + round(sf_row))

        if "INDEX_REVERSED" in flags and to_int(r[7]) <= 0:
            continue

        lines.append(bill_sql(period_code, customer_code, cons_excel, up_row, sf_row, elec, total))
        period_bill += 1

    period_stats.append((period_code, period_name, len(rows), period_mr, period_bill))
    total_mr   += period_mr
    total_bill += period_bill


# ─────────────────────────────────────────────────────────────
# Assemble file
# ─────────────────────────────────────────────────────────────

header = [
    "-- " + "═" * 62,
    "-- V13__import_multi_month_periods.sql",
    "-- Generated by scripts/generate_v13_multi_month.py",
    f"-- Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "--",
    "-- Treats each multi-month sheet as a single billing period.",
    f"-- Periods:       {len(period_stats)}",
    f"-- Meter readings: {total_mr}",
    f"-- Bills:          {total_bill}",
    "--",
    "-- Depends on: V10, V11, V12 (must be applied first)",
    "-- " + "═" * 62,
    "",
    "SET NAMES utf8mb4;",
    "SET FOREIGN_KEY_CHECKS = 0;",
]

footer = ["", "SET FOREIGN_KEY_CHECKS = 1;", ""]

content = "\n".join(header + lines + footer)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write(content)

print()
print("=" * 58)
print("  V13 Summary")
print("=" * 58)
for (code, name, customers, mr, bill) in period_stats:
    print(f"  {code}")
    print(f"    Name: {name}")
    print(f"    Customers: {customers}  |  MR: {mr}  |  Bills: {bill}")
print()
print(f"  Total meter readings: {total_mr}")
print(f"  Total bills:          {total_bill}")
print(f"  Rows skipped:         {skipped_rows}")
print()
print(f"  Output: {OUTPUT_PATH}")
print(f"  File size: {len(content.encode('utf-8')):,} bytes")
print("=" * 58)
