#!/usr/bin/env python3
"""
generate_v12_need_review.py

Generates V12__import_need_review.sql which imports the remaining valid rows
from the need_review sheet that were excluded from V10:

  - INDEX_REVERSED (meter replacement):  prev=0, curr=cons_excel
  - MISSING_INDEX with service_fee:      import with 0 or actual consumption
  - CONSUMPTION_MISMATCH (non-reversed): import with cons_excel for bill
  - MULTI_MONTH_ROW / no service_fee:    skipped

Run AFTER V11 has been applied.

Usage:
    python3 scripts/generate_v12_need_review.py
"""

import os, calendar
from datetime import date, datetime
from collections import Counter, defaultdict

try:
    import openpyxl
except ImportError:
    raise SystemExit("Run: python3 -m pip install openpyxl")

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

EXCEL_PATH   = os.path.join(BASE_DIR, "docs", "data",
               "clean_data_migration_workbook2020_2026.xlsx")
OUTPUT_PATH  = os.path.join(BASE_DIR, "src", "main", "resources", "db", "migration",
               "V12__import_need_review.sql")

EXCLUDED_CUSTOMERS = {"KH093"}


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def sql_str(v):
    if v is None: return "NULL"
    return "'" + str(v).replace("\\", "\\\\").replace("'", "\\'") + "'"

def to_int(v, default=0):
    if v is None: return default
    try: return int(float(v))
    except: return default

def to_float(v, default=0.0):
    if v is None: return default
    try: return float(v)
    except: return default

def round0(v): return round(to_float(v))

def is_valid_period(code):
    if not code or "_to_" in str(code): return False
    s = str(code)
    if len(s) != 7 or s[4] != "-": return False
    try:
        y, m = int(s[:4]), int(s[5:7])
        return 2000 <= y <= 2030 and 1 <= m <= 12
    except: return False

def period_end(code):
    y, m = int(code[:4]), int(code[5:7])
    last = calendar.monthrange(y, m)[1]
    return date(y, m, last)


# ─────────────────────────────────────────────────────────────
# Read Excel
# ─────────────────────────────────────────────────────────────

print("Reading Excel …")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

# Collect unit_prices already known per period (from import_candidates)
ws_ic = wb["import_candidates"]
ic_rows = list(ws_ic.iter_rows(values_only=True))[1:]
period_unit_prices = defaultdict(list)
period_service_fees = defaultdict(list)
for r in ic_rows:
    if r[9] is not None: period_unit_prices[r[0]].append(to_float(r[9]))
    if r[11] is not None: period_service_fees[r[0]].append(to_float(r[11]))
period_mode_up = {p: Counter(ups).most_common(1)[0][0] for p, ups in period_unit_prices.items()}
period_mode_sf = {p: Counter(sfs).most_common(1)[0][0] for p, sfs in period_service_fees.items()}

# Read need_review
ws_nr = wb["need_review"]
nr_data = list(ws_nr.iter_rows(values_only=True))[1:]
wb.close()
print(f"  need_review rows: {len(nr_data)}")


# ─────────────────────────────────────────────────────────────
# Categorize rows
# ─────────────────────────────────────────────────────────────

rows_meter_reset  = []   # INDEX_REVERSED → meter replacement
rows_missing_idx  = []   # MISSING_INDEX with service_fee
rows_mismatch     = []   # CONSUMPTION_MISMATCH (non-reversed)
skipped           = Counter()

for r in nr_data:
    period_code   = r[0]
    customer_code = r[1]
    flags_str     = str(r[15]) if r[15] else ""
    flags         = {f.strip() for f in flags_str.split(";")} if flags_str else set()

    # Hard skips
    if "MULTI_MONTH_ROW" in flags:
        skipped["MULTI_MONTH_ROW"] += 1
        continue
    if not is_valid_period(period_code):
        skipped["invalid_period"] += 1
        continue
    if customer_code in EXCLUDED_CUSTOMERS:
        skipped["KH093"] += 1
        continue

    if "INDEX_REVERSED" in flags:
        # Meter replacement: prev=0, curr=cons_excel
        cons = to_int(r[7])            # cons_excel = consumption on new meter
        if cons <= 0:
            skipped["INDEX_REVERSED_zero_cons"] += 1
            continue
        up_val  = to_float(r[9]) or period_mode_up.get(period_code, 0)
        sf_val  = to_float(r[11]) if r[11] is not None else period_mode_sf.get(period_code, 20000)
        elec_v  = round0(r[10]) if r[10] is not None else round(cons * up_val)
        total_v = round0(r[12]) if r[12] is not None else (elec_v + round(sf_val))
        rows_meter_reset.append({
            "period_code":   period_code,
            "customer_code": customer_code,
            "prev_index":    0,
            "curr_index":    cons,
            "consumption":   cons,
            "unit_price":    up_val,
            "service_fee":   sf_val,
            "elec_amount":   elec_v,
            "total_amount":  total_v,
        })

    elif "MISSING_INDEX" in flags:
        # Only import if service_fee is known (row had a bill)
        sf = to_float(r[11]) if r[11] is not None else period_mode_sf.get(period_code)
        total = r[12]
        if sf is None and total is None:
            skipped["MISSING_INDEX_no_sf"] += 1
            continue
        # curr may be 0 (genuinely no reading) or a small number
        prev = to_int(r[5], 0)
        curr = to_int(r[6], prev)      # fallback = same as prev
        if curr < prev:
            skipped["MISSING_INDEX_reversed"] += 1
            continue
        cons     = curr - prev
        elec     = round0(r[10]) if r[10] is not None else round(cons * (to_float(r[9]) or period_mode_up.get(period_code, 0)))
        total_v  = round0(total) if total is not None else (elec + round(sf))
        rows_missing_idx.append({
            "period_code":   period_code,
            "customer_code": customer_code,
            "prev_index":    prev,
            "curr_index":    curr,
            "consumption":   cons,
            "unit_price":    to_float(r[9]) or period_mode_up.get(period_code, 0),
            "service_fee":   sf,
            "elec_amount":   elec,
            "total_amount":  total_v,
        })

    elif "CONSUMPTION_MISMATCH" in flags:
        # Trust cons_excel for billing; indices inserted as-is
        prev = to_int(r[5], 0)
        curr = to_int(r[6], prev)
        if curr < prev:
            skipped["MISMATCH_reversed"] += 1
            continue
        cons = to_int(r[7])
        sf   = to_float(r[11]) if r[11] is not None else period_mode_sf.get(period_code, 20000)
        up   = to_float(r[9]) or period_mode_up.get(period_code, 0)
        elec = round0(r[10]) if r[10] is not None else round(cons * up)
        tot  = round0(r[12]) if r[12] is not None else (elec + round(sf))
        rows_mismatch.append({
            "period_code":   period_code,
            "customer_code": customer_code,
            "prev_index":    prev,
            "curr_index":    curr,
            "consumption":   cons,
            "unit_price":    up,
            "service_fee":   sf,
            "elec_amount":   elec,
            "total_amount":  tot,
        })
    else:
        skipped["other"] += 1

all_rows = rows_meter_reset + rows_missing_idx + rows_mismatch
print(f"  INDEX_REVERSED (meter reset):  {len(rows_meter_reset)}")
print(f"  MISSING_INDEX with billing:    {len(rows_missing_idx)}")
print(f"  CONSUMPTION_MISMATCH:          {len(rows_mismatch)}")
print(f"  Total to import:               {len(all_rows)}")
print(f"  Skipped: {dict(skipped)}")


# ─────────────────────────────────────────────────────────────
# Generate SQL helper
# ─────────────────────────────────────────────────────────────

def mr_sql(row):
    p  = row["period_code"]
    c  = row["customer_code"]
    ed = period_end(p)
    read_at = f"{ed} 12:00:00"
    return (
        f"INSERT IGNORE INTO meter_reading "
        f"(period_id, customer_id, previous_index, current_index, read_at, created_at, updated_at) "
        f"SELECT bp.id, c.id, {row['prev_index']}, {row['curr_index']}, "
        f"'{read_at}', NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(p)} AND c.code = {sql_str(c)};"
    )

def bill_sql(row):
    p          = row["period_code"]
    c          = row["customer_code"]
    cons       = row["consumption"]
    up         = row["unit_price"]
    sf         = row["service_fee"]
    elec       = row["elec_amount"]
    svc        = round(sf)
    total      = row["total_amount"]
    payment_c  = f"TIENDIEN {p} {c}"
    return (
        f"INSERT IGNORE INTO bill "
        f"(period_id, customer_id, consumption, unit_price, service_fee, "
        f"electricity_amount, service_amount, total_amount, paid_amount, "
        f"status, payment_code, sent_via_zalo, created_at, updated_at) "
        f"SELECT bp.id, c.id, {cons}, {up}, {sf}, "
        f"{elec}, {svc}, {total}, {total}, "
        f"'PAID', {sql_str(payment_c)}, FALSE, NOW(), NOW() "
        f"FROM billing_period bp, customer c "
        f"WHERE bp.code = {sql_str(p)} AND c.code = {sql_str(c)};"
    )


# ─────────────────────────────────────────────────────────────
# Build SQL output
# ─────────────────────────────────────────────────────────────

lines = [
    "-- " + "═" * 62,
    "-- V12__import_need_review.sql",
    "-- Generated by scripts/generate_v12_need_review.py",
    f"-- Generated at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
    "--",
    "-- Imports rows from need_review sheet that were excluded from V10:",
    f"--   INDEX_REVERSED (meter replacement): {len(rows_meter_reset)} rows",
    f"--   MISSING_INDEX with billing data:    {len(rows_missing_idx)} rows",
    f"--   CONSUMPTION_MISMATCH:               {len(rows_mismatch)} rows",
    "--",
    "-- Depends on: V10, V11 (must be applied first)",
    "-- " + "═" * 62,
    "",
    "SET NAMES utf8mb4;",
    "SET FOREIGN_KEY_CHECKS = 0;",
]

# ── Group A: Meter resets ──────────────────────────────────
lines += [
    "",
    "-- " + "─" * 62,
    f"-- GROUP A: INDEX_REVERSED — meter replacement ({len(rows_meter_reset)} rows)",
    "-- previous_index set to 0 (new meter); current_index = consumption on new meter",
    "-- " + "─" * 62,
    "-- Meter readings",
]
for row in rows_meter_reset:
    lines.append(mr_sql(row))

lines += ["", "-- Bills"]
for row in rows_meter_reset:
    lines.append(bill_sql(row))

# ── Group B: Missing index (with billing) ─────────────────
lines += [
    "",
    "-- " + "─" * 62,
    f"-- GROUP B: MISSING_INDEX with service_fee ({len(rows_missing_idx)} rows)",
    "-- Customers who had 0 or unknown consumption but were still charged service fee",
    "-- " + "─" * 62,
    "-- Meter readings",
]
for row in rows_missing_idx:
    lines.append(mr_sql(row))

lines += ["", "-- Bills"]
for row in rows_missing_idx:
    lines.append(bill_sql(row))

# ── Group C: Consumption mismatch ─────────────────────────
lines += [
    "",
    "-- " + "─" * 62,
    f"-- GROUP C: CONSUMPTION_MISMATCH (non-reversed) ({len(rows_mismatch)} rows)",
    "-- Indices inserted as-is; bill.consumption uses cons_excel (what was actually billed)",
    "-- " + "─" * 62,
    "-- Meter readings",
]
for row in rows_mismatch:
    lines.append(mr_sql(row))

lines += ["", "-- Bills"]
for row in rows_mismatch:
    lines.append(bill_sql(row))

lines += ["", "SET FOREIGN_KEY_CHECKS = 1;", ""]

content = "\n".join(lines)
with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write(content)

print()
print("=" * 55)
print("  V12 Summary")
print("=" * 55)
print(f"  Meter resets (INDEX_REVERSED): {len(rows_meter_reset)}")
print(f"  Missing index (with billing):  {len(rows_missing_idx)}")
print(f"  Consumption mismatch:          {len(rows_mismatch)}")
print(f"  Total rows:                    {len(all_rows)}")
print(f"  Skipped:                       {sum(skipped.values())}")
for reason, cnt in skipped.items():
    print(f"    {reason}: {cnt}")
print()
print(f"  Output: {OUTPUT_PATH}")
print(f"  File size: {len(content.encode('utf-8')):,} bytes")
print("=" * 55)
